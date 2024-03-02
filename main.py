import openpyxl
import pandas as pd
import xlsxwriter
import matplotlib.pyplot as plt
#
workbook = openpyxl.load_workbook('DuLieuThucHanh2_K66_GuiSV_V1.xlsx')
#Thêm data sheet 1
sheet = workbook['Sinh Vien']
new_data = [
    ['4', '0206566', 'Nguyen Truong Toan', '2003-05-10', 'Nam', '66PM3'],
    ['5', '0192322', 'Nguyen Huu Thang', '2003-02-22', 'Nam', '66PM3'],
    ['6', '0182344', 'Vu Van Thao', '2003-12-14', 'Nam', '66PM3'],
    ['7', '123789', 'Nguyen Kien Cuong', '2003-08-30', 'Nam', '66PM3'],
    ['8', '890123', 'Bui Ngoc Tien', '2003-06-25', 'Nam', '66PM3'],
    ['9', '456789', 'Nguyen Xuan Tai', '2003-04-03', 'Nam', '66PM3'],
    ['10', '234567', 'Bui Van K', '2003-10-15', 'Nam', '66PM3'],
    ['11', '678901', 'Vu Thi L', '2003-11-28', 'Nu', '66PM3'],
    ['12', '345678', 'Pham Van M', '2003-09-20', 'Nam', '66PM3'],
    ['13', '901234', 'Le Van N', '2003-03-05', 'Nam', '66PM3'],
    ['14', '567890', 'Tran Thi O', '2003-01-08', 'Nu', '66PM3'],
    ['15', '123789', 'Nguyen Van P', '2003-07-17', 'Nam', '66PM3'],
    ['16', '890123', 'Hoang Van Q', '2003-04-12', 'Nam', '66PM3'],
    ['17', '456789', 'Dang Van R', '2003-10-09', 'Nam', '66PM3']
]
for row_index, row_data in enumerate(new_data, start=5):  # Bắt đầu từ hàng 5
    for col_index, cell_data in enumerate(row_data, start=1):
        sheet.cell(row=row_index, column=col_index, value=cell_data)

sheet = workbook['Diem Mon 1']
new_data = [
    ['4', '0206566', 'Nguyen Truong Toan', '8.5', '9.0'],
    ['5', '0192322', 'Nguyen Huu Thang', '7.0', '8.0'],
    ['6', '0182344', 'Vu Van Thao', '4.0', '9.5'],
    ['7', '123789', 'Nguyen Kien Cuong', '8.5', '4.0'],
    ['8', '890123', 'Bui Ngoc Tien', '7.0', '2.5'],
    ['9', '456789', 'Nguyen Xuan Tai', '6.0', '2.0'],
    ['10', '234567', 'Bui Van K', '7.5', '2.0'],
    ['11', '678901', 'Vu Thi L', '0', '4.0'],
    ['12', '345678', 'Pham Van M', '2.5', '8.0'],
    ['13', '901234', 'Le Van N', '10.0', '7.0'],
    ['14', '567890', 'Tran Thi O', '2.5', '4.0'],
    ['15', '123789', 'Nguyen Van P', '5.0', '5.0'],
    ['16', '890123', 'Hoang Van Q', '7.5', '7.0'],
    ['17', '456789', 'Dang Van R', '6.0', '6.5']
]
for row_index, row_data in enumerate(new_data, start=5):  # Bắt đầu từ hàng 5
    for col_index, cell_data in enumerate(row_data, start=1):
        sheet.cell(row=row_index, column=col_index, value=cell_data)

new_data = [
    ['4', '0206566', 'Nguyen Truong Toan', '6.5', '7.0'],
    ['5', '0192322', 'Nguyen Huu Thang', '8.0', '9.0'],
    ['6', '0182344', 'Vu Van Thao', '5.0', '8.5'],
    ['7', '123789', 'Nguyen Kien Cuong', '9.0', '6.0'],
    ['8', '890123', 'Bui Ngoc Tien', '6.5', '3.0'],
    ['9', '456789', 'Nguyen Xuan Tai', '7.0', '4.0'],
    ['10', '234567', 'Bui Van K', '6.0', '3.5'],
    ['11', '678901', 'Vu Thi L', '3.5', '5.0'],
    ['12', '345678', 'Pham Van M', '7.0', '7.5'],
    ['13', '901234', 'Le Van N', '8.5', '6.0'],
    ['14', '567890', 'Tran Thi O', '4.0', '5.5'],
    ['15', '123789', 'Nguyen Van P', '6.0', '6.0'],
    ['16', '890123', 'Hoang Van Q', '8.0', '8.5'],
    ['17', '456789', 'Dang Van R', '7.0', '7.0']
]

sheet = workbook['Diem Mon 2']

for row_index, row_data in enumerate(new_data, start=5):  # Bắt đầu từ hàng 5
    for col_index, cell_data in enumerate(row_data, start=1):
        sheet.cell(row=row_index, column=col_index, value=cell_data)

new_data = [
    ['4', '0206566', 'Nguyen Truong Toan', '9.0', '8.0'],
    ['5', '0192322', 'Nguyen Huu Thang', '6.5', '7.5'],
    ['6', '0182344', 'Vu Van Thao', '8.0', '9.0'],
    ['7', '123789', 'Nguyen Kien Cuong', '7.5', '8.5'],
    ['8', '890123', 'Bui Ngoc Tien', '6.0', '5.0'],
    ['9', '456789', 'Nguyen Xuan Tai', '5.5', '6.0'],
    ['10', '234567', 'Bui Van K', '8.5', '9.0'],
    ['11', '678901', 'Vu Thi L', '4.0', '5.5'],
    ['12', '345678', 'Pham Van M', '9.0', '8.5'],
    ['13', '901234', 'Le Van N', '7.0', '7.5'],
    ['14', '567890', 'Tran Thi O', '6.0', '6.0'],
    ['15', '123789', 'Nguyen Van P', '8.0', '8.0'],
    ['16', '890123', 'Hoang Van Q', '5.0', '4.5'],
    ['17', '456789', 'Dang Van R', '6.5', '7.0']
]
sheet = workbook['Diem Mon 3']
for row_index, row_data in enumerate(new_data, start=5):  # Bắt đầu từ hàng 5
    for col_index, cell_data in enumerate(row_data, start=1):
        sheet.cell(row=row_index, column=col_index, value=cell_data)

workbook.save('DuLieuThucHanh2_K66_GuiSV_V1.xlsx')

df_sinh_vien = pd.DataFrame(workbook['Sinh Vien'].values)
df_diem_mon_1 = pd.DataFrame(workbook['Diem Mon 1'].values)
df_diem_mon_2 = pd.DataFrame(workbook['Diem Mon 2'].values)
df_diem_mon_3 = pd.DataFrame(workbook['Diem Mon 3'].values)

print("DataFrame Sinh Vien:")
print(df_sinh_vien)
print("\nDataFrame Diem Mon 1:")
print(df_diem_mon_1)
print("\nDataFrame Diem Mon 2:")
print(df_diem_mon_2)
print("\nDataFrame Diem Mon 3:")
print(df_diem_mon_3)

#Câu c
# Tạo DataFrame df_tong_hop từ df_sinh_vien
df_tong_hop = df_sinh_vien.copy()

# Đổi tên cột cho df_tong_hop
df_tong_hop.columns = ['STT', 'Ma Sinh Vien', 'Ho Ten', 'Ngay Sinh', 'Gioi Tinh', 'Lop Quan Ly']

# Lấy điểm cuối kỳ từ cột thứ 4 của các DataFrame df_diem_mon_i
df_tong_hop['Diem Thi Mon 1'] = df_diem_mon_1.iloc[1:, 4].astype(float)
df_tong_hop['Diem Thi Mon 2'] = df_diem_mon_2.iloc[1:, 4].astype(float)
df_tong_hop['Diem Thi Mon 3'] = df_diem_mon_3.iloc[1:, 4].astype(float)

# Tính điểm tổng kết cho từng môn
df_tong_hop['Diem Tong Ket Mon 1'] = 0.3 * df_diem_mon_1.iloc[1:, 3].astype(float) + 0.7 * df_diem_mon_1.iloc[1:, 4].astype(float)
df_tong_hop['Diem Tong Ket Mon 2'] = 0.3 * df_diem_mon_2.iloc[1:, 3].astype(float) + 0.7 * df_diem_mon_2.iloc[1:, 4].astype(float)
df_tong_hop['Diem Tong Ket Mon 3'] = 0.3 * df_diem_mon_3.iloc[1:, 3].astype(float) + 0.7 * df_diem_mon_3.iloc[1:, 4].astype(float)
df_tong_hop['Diem Trung Binh'] = (df_tong_hop['Diem Tong Ket Mon 1'] + df_tong_hop['Diem Tong Ket Mon 2'] + df_tong_hop['Diem Tong Ket Mon 3']) / 3


# cau d
df = df_tong_hop

df_khong_qua_mon = df[(df['Diem Tong Ket Mon 1'] < 4.0) |
                      (df['Diem Tong Ket Mon 2'] < 4.0) |
                      (df['Diem Tong Ket Mon 3'] < 4.0)]

# In ra danh sách sinh viên không qua môn
print("Danh sách sinh viên không qua môn:")
print(df_khong_qua_mon[['Ma Sinh Vien', 'Ho Ten', 'Diem Tong Ket Mon 1', 'Diem Tong Ket Mon 2', 'Diem Tong Ket Mon 3']])


# cau e

df_sv_khong_truot = df_tong_hop[(df_tong_hop['Diem Tong Ket Mon 1'] > 4) &
                                   (df_tong_hop['Diem Tong Ket Mon 2'] > 4) &
                                   (df_tong_hop['Diem Tong Ket Mon 3'] > 4)]
# cau f


sinh_vien_labels = df_tong_hop['Ho Ten']
diem_trung_binh = df_tong_hop['Diem Trung Binh']

plt.figure(figsize=(10, 6))
plt.bar(sinh_vien_labels, diem_trung_binh, color='blue')
plt.xlabel('Họ và Tên Sinh Viên')
plt.ylabel('Điểm Trung Bình')
plt.title('Kết Quả Học Tập Của Sinh Viên')
plt.xticks(rotation=45, ha='right')
plt.tight_layout()
plt.show()

# xuất ra sheet trong excel

existing_excel = pd.ExcelFile('DuLieuThucHanh2_K66_GuiSV_V1.xlsx')
existing_sheets = existing_excel.sheet_names

# Mở workbook để thêm dữ liệu mới mà không thay đổi dữ liệu đã tồn tại
with pd.ExcelWriter('DuLieuThucHanh2_K66_GuiSV_V1.xlsx', engine='openpyxl', mode='a') as writer:
    # Ghi các DataFrame mới vào các sheet mới
    df_tong_hop.to_excel(writer, sheet_name='Tong hop', index=False)
    df_khong_qua_mon.to_excel(writer, sheet_name='Sinh vien khong qua mon', index=False)
    df_sv_khong_truot.to_excel(writer, sheet_name='Sinh vien khong truot mon nao', index=False)









